"""Run the frozen Excel golden set against a candidate model and keep all rows."""

from __future__ import annotations

import argparse
from datetime import UTC, datetime
import json
from pathlib import Path
import re
from typing import Any

import joblib

from .constants import SIGNAL_CODES
from .guidance import is_victim_recovery_request
from .local_rules import evaluate_local_rules, load_rule_bundle
from .model import PhishingSignalModel
from .redact import redact_l2


def _expected_risks(expected: str) -> set[str]:
    lowered = expected.lower()
    if "risk:" not in lowered:
        return set()
    tail = lowered.split("risk:", 1)[1].split(".", 1)[0]
    return {risk for risk in ("high", "medium", "unknown") if risk in tail}


def _score_bounds(expected: str) -> tuple[float | None, float | None]:
    minimum: float | None = None
    maximum: float | None = None
    if match := re.search(r"score\s*>=\s*(0(?:\.\d+)?|1(?:\.0+)?)", expected, re.I):
        minimum = float(match.group(1))
    if match := re.search(r"score\s*<\s*(0(?:\.\d+)?|1(?:\.0+)?)", expected, re.I):
        maximum = float(match.group(1))
    if match := re.search(
        r"score\s+in\s*\[\s*(0(?:\.\d+)?)\s*,\s*(0(?:\.\d+)?)\s*\]",
        expected,
        re.I,
    ):
        minimum, maximum = float(match.group(1)), float(match.group(2))
    return minimum, maximum


def _required_signals(expected: str) -> set[str]:
    if re.search(r"signals\s*:\s*\[\s*\]", expected, re.I):
        return set()
    return {code for code in SIGNAL_CODES if code in expected}


def _predict(model: PhishingSignalModel, text: str, bundle: dict) -> dict[str, Any]:
    local = evaluate_local_rules(text, bundle)
    redaction = redact_l2(text)
    if local.otp_blocked or redaction.otp_found:
        return {
            "risk": "high",
            "score": 1.0,
            "signals": [{"code": "yeu_cau_otp", "confidence": 1.0, "evidence": ""}],
            "explanation": "Tin nhắn chứa hoặc yêu cầu mã xác nhận. Đừng chia sẻ mã.",
            "local_signals": list(local.local_signals),
            "redacted_text": redaction.text,
        }
    gate = bundle["l1"]["gate"]
    local_rules = bundle["l1"]["local_signals"]
    local_score = sum(
        max(0.0, float(local_rules[name].get("boost", 0.0)))
        for name in local.local_signals
    )
    always_call = any(
        name in gate["always_call_when_local_signal"]
        for name in local.local_signals
    )
    if (
        len(local.normalized) < int(gate["min_length_to_call_server"])
        or (
            not always_call
            and local_score < float(gate["min_score_to_call_server"])
        )
    ):
        return {
            "risk": "unknown",
            "score": 0.0,
            "signals": [],
            "explanation": "Chưa đủ thông tin để kết luận là an toàn hay lừa đảo.",
            "local_signals": list(local.local_signals),
            "redacted_text": redaction.text,
        }
    prediction = model.predict(
        redaction.text,
        signal_boosts=local.signal_boosts,
    )
    return {
        **prediction,
        "local_signals": list(local.local_signals),
        "redacted_text": redaction.text,
    }


def _evaluate_case(
    model: PhishingSignalModel,
    *,
    case_id: str,
    text: str,
    expected: str,
    bundle: dict,
) -> dict[str, Any]:
    prediction = _predict(model, text, bundle)
    actual_signals = {
        str(item["code"]) for item in prediction.get("signals", [])
    }
    expected_risks = _expected_risks(expected)
    required_signals = _required_signals(expected)
    minimum, maximum = _score_bounds(expected)
    failures: list[str] = []
    if expected_risks and prediction["risk"] not in expected_risks:
        failures.append(
            f"risk={prediction['risk']} not in {sorted(expected_risks)}"
        )
    if not required_signals.issubset(actual_signals):
        failures.append(
            f"missing_signals={sorted(required_signals - actual_signals)}"
        )
    score = float(prediction["score"])
    if minimum is not None and score < minimum:
        failures.append(f"score={score:.6f} < {minimum}")
    if maximum is not None and score >= maximum:
        failures.append(f"score={score:.6f} >= {maximum}")
    if "Out-of-scope" in expected:
        if is_victim_recovery_request(text):
            explanation = str(prediction.get("explanation", "")).lower()
            if "ngân hàng" not in explanation or "công an" not in explanation:
                failures.append("missing recovery guidance")
        elif prediction["risk"] not in {"medium", "high"}:
            failures.append("investment warning was not raised")
    return {
        "id": case_id,
        "input": text,
        "expected": expected,
        "actual_risk": prediction["risk"],
        "actual_score": score,
        "actual_signals": sorted(actual_signals),
        "local_signals": prediction["local_signals"],
        "explanation": prediction.get("explanation", ""),
        "passed": not failures,
        "failures": failures,
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Evaluate the 20-case CHAN Excel golden set."
    )
    parser.add_argument("--model", type=Path, required=True)
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--rules", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--sheet", default="Golden Set (20 Cases)")
    return parser


def main() -> None:
    args = build_parser().parse_args()
    try:
        import openpyxl
    except ImportError as error:
        raise RuntimeError(
            "openpyxl is required; install codebase/ml[workbook]"
        ) from error

    model = joblib.load(args.model)
    if not isinstance(model, PhishingSignalModel):
        raise TypeError("artifact is not a PhishingSignalModel")
    workbook = openpyxl.load_workbook(
        args.workbook,
        data_only=True,
        read_only=True,
    )
    if args.sheet not in workbook.sheetnames:
        raise ValueError(f"missing workbook sheet: {args.sheet}")
    bundle = load_rule_bundle(args.rules)
    cases: list[dict[str, Any]] = []
    for row in workbook[args.sheet].iter_rows(min_row=2, values_only=True):
        case_id, text, expected = row[0], row[3], row[4]
        if not case_id or not text or not expected:
            continue
        cases.append(
            _evaluate_case(
                model,
                case_id=str(case_id),
                text=str(text),
                expected=str(expected),
                bundle=bundle,
            )
        )
    passed = sum(bool(case["passed"]) for case in cases)
    payload = {
        "evaluated_at": datetime.now(UTC).isoformat(),
        "model": args.model.name,
        "rule_bundle_version": bundle.get("bundle_version"),
        "pass_count": passed,
        "total_count": len(cases),
        "result": f"{passed}/{len(cases)}",
        "cases": cases,
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
